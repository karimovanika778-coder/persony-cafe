from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from database import db, Workshop, Product, Stock, Movement, Order, Recipe, RecipeIngredient
from datetime import datetime
import os
#import cv2
import numpy as np
from pyzbar.pyzbar import decode
import threading
import time

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'  # в реальном проекте меняем!
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cafe_inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Создаем таблицы при первом запуске
with app.app_context():
    db.create_all()

    # Добавляем тестовые данные, если база пустая
    if Workshop.query.count() == 0:
        # Создаем цеха
        workshops = [
            Workshop(name='Горячий цех', description='Горячие блюда, супы, гарниры'),
            Workshop(name='Холодный цех', description='Салаты, закуски, холодные напитки'),
            Workshop(name='Мясной цех', description='Разделка мяса, полуфабрикаты'),
            Workshop(name='Кондитерский цех', description='Десерты, выпечка'),
            Workshop(name='Открытая кухня', description='Блюда на виду у гостей'),
            Workshop(name='Заморозка', description='Замороженные продукты')
        ]
        db.session.add_all(workshops)

        # Создаем тестовые продукты
        products = [
            Product(name='Вырезка говяжья', unit='кг', category='Мясо', type='ingredient'),
            Product(name='Куриное филе', unit='кг', category='Птица', type='ingredient'),
            Product(name='Лосось слабосоленый (привозной)', unit='кг', category='Рыба', type='retail'),
            Product(name='Чизкейк готовый', unit='шт', category='Десерты', type='ready_dish'),
            Product(name='Сливочное масло', unit='кг', category='Молочка', type='ingredient'),
        ]
        db.session.add_all(products)
        db.session.commit()

        # Добавляем начальные остатки (для примера)
        stock1 = Stock(product_id=1, workshop_id=3, quantity=5.5)  # 5.5 кг говядины в мясном
        stock2 = Stock(product_id=2, workshop_id=3, quantity=3.0)  # 3 кг курицы в мясном
        stock3 = Stock(product_id=3, workshop_id=2, quantity=2.0)  # 2 кг лосося в холодном
        stock4 = Stock(product_id=4, workshop_id=4, quantity=4.0)  # 4 чизкейка в кондитерском
        stock5 = Stock(product_id=5, workshop_id=5, quantity=1.2)  # 1.2 кг масла на открытой кухне

        db.session.add_all([stock1, stock2, stock3, stock4, stock5])
        db.session.commit()


# Главная страница - выбор цеха
@app.route('/')
def index():
    workshops = Workshop.query.all()
    return render_template('index.html', workshops=workshops)


# Страница цеха (инвентаризация)
@app.route('/workshop/<int:workshop_id>')
def workshop(workshop_id):
    workshop = Workshop.query.get_or_404(workshop_id)
    # Получаем все продукты с остатками в этом цехе
    stocks = Stock.query.filter_by(workshop_id=workshop_id).all()
    return render_template('workshop.html', workshop=workshop, stocks=stocks)


# Поиск продуктов по всем цехам
@app.route('/search')
def search():
    query = request.args.get('q', '')
    if query:
        # Ищем продукты по названию
        products = Product.query.filter(Product.name.contains(query)).all()

        # Собираем информацию о наличии во всех цехах
        results = []
        for product in products:
            stocks = Stock.query.filter_by(product_id=product.id).all()
            results.append({
                'product': product,
                'stocks': stocks
            })
    else:
        results = []

    return render_template('search.html', query=query, results=results)


# Обновление количества продукта в цехе
@app.route('/update_stock', methods=['POST'])
def update_stock():
    product_id = request.form.get('product_id')
    workshop_id = request.form.get('workshop_id')
    new_quantity = request.form.get('quantity', type=float)

    stock = Stock.query.filter_by(
        product_id=product_id,
        workshop_id=workshop_id
    ).first()

    if stock:
        stock.quantity = new_quantity
        flash(f'Количество обновлено', 'success')
    else:
        # Если такой записи нет, создаем новую
        stock = Stock(
            product_id=product_id,
            workshop_id=workshop_id,
            quantity=new_quantity
        )
        db.session.add(stock)
        flash(f'Продукт добавлен в цех', 'success')

    db.session.commit()
    return redirect(url_for('workshop', workshop_id=workshop_id))


# Добавление нового продукта (с автоматическим объединением)
@app.route('/add_product', methods=['GET', 'POST'])
def add_product():
    if request.method == 'POST':
        name = request.form.get('name')
        barcode = request.form.get('barcode')
        unit = request.form.get('unit')
        category = request.form.get('category')
        product_type = request.form.get('type')

        # Проверяем, есть ли уже продукт с таким названием (регистронезависимо)
        existing_product = Product.query.filter(
            Product.name.ilike(name.strip())  # ilike - регистронезависимый поиск
        ).first()

        if existing_product:
            # Продукт уже существует - предлагаем объединить
            flash(f'⚠️ Продукт "{name}" уже существует в базе!', 'warning')

            # Если есть штрихкод и у существующего продукта его нет - добавляем
            if barcode and not existing_product.barcode:
                existing_product.barcode = barcode
                db.session.commit()
                flash(f'✅ Штрихкод добавлен к существующему продукту', 'success')

            # Если штрихкод есть и у существующего продукта другой - предупреждаем
            elif barcode and existing_product.barcode and existing_product.barcode != barcode:
                flash(f'⚠️ У продукта уже есть другой штрихкод: {existing_product.barcode}', 'warning')

            # Если категория не указана, но есть у существующего - копируем
            if not category and existing_product.category:
                category = existing_product.category
                flash(f'📋 Категория скопирована: {category}', 'info')

            # Если тип не указан, но есть у существующего - копируем
            if not product_type and existing_product.type:
                product_type = existing_product.type
                flash(f'📋 Тип скопирован: {product_type}', 'info')

            # Создаём новый продукт? Нет, предлагаем использовать существующий
            flash(f'💡 Используйте существующий продукт: {existing_product.name}', 'info')

            # Перенаправляем на страницу добавления с предзаполненными данными
            return redirect(url_for('add_product',
                                    name=existing_product.name,
                                    barcode=existing_product.barcode,
                                    unit=existing_product.unit,
                                    category=existing_product.category,
                                    type=existing_product.type))
        else:
            # Создаем новый продукт
            product = Product(
                name=name,
                barcode=barcode,
                unit=unit,
                category=category,
                type=product_type
            )
            db.session.add(product)
            db.session.commit()

            flash(f'✅ Продукт {name} добавлен', 'success')
            return redirect(url_for('index'))

    # GET запрос - показываем форму
    # Если есть параметры в URL, предзаполняем форму
    name = request.args.get('name', '')
    barcode = request.args.get('barcode', '')
    unit = request.args.get('unit', 'кг')
    category = request.args.get('category', '')
    product_type = request.args.get('type', 'ingredient')

    return render_template('add_product.html',
                           name=name,
                           barcode=barcode,
                           unit=unit,
                           category=category,
                           product_type=product_type)


# Сканер штрихкодов
@app.route('/scan')
def scan():
    return render_template('scan.html')


@app.route('/scan_barcode')
def scan_barcode_route():
    try:
        from scanner import scan_barcode
        barcode = scan_barcode()

        if barcode:
            # Ищем продукт по штрихкоду
            product = Product.query.filter_by(barcode=barcode).first()

            if product:
                flash(f'🔍 Найден продукт: {product.name}', 'success')
                return redirect(url_for('workshop', workshop_id=1))
            else:
                # Ищем продукт, который мог быть добавлен без штрихкода
                # Предлагаем пользователю ввести название
                flash(f'📦 Штрихкод {barcode} не найден. Введите название продукта.', 'warning')
                return redirect(url_for('add_product', barcode=barcode))
        else:
            flash('Сканирование отменено', 'info')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Ошибка сканера: {str(e)}', 'error')
        return redirect(url_for('index'))


# Страница прихода/расхода для продукта
@app.route('/movement/<int:product_id>/<int:workshop_id>')
def movement(product_id, workshop_id):
    product = Product.query.get_or_404(product_id)
    workshop = Workshop.query.get_or_404(workshop_id)
    stock = Stock.query.filter_by(product_id=product_id, workshop_id=workshop_id).first()
    current_quantity = stock.quantity if stock else 0

    return render_template('movement.html',
                           product=product,
                           workshop=workshop,
                           current_quantity=current_quantity)


# Обработка прихода
@app.route('/income', methods=['POST'])
def income():
    product_id = request.form.get('product_id')
    workshop_id = request.form.get('workshop_id')
    quantity = request.form.get('quantity', type=float)
    reason = request.form.get('reason', 'поставка')
    comment = request.form.get('comment', '')

    if quantity <= 0:
        flash('Количество должно быть положительным', 'error')
        return redirect(url_for('movement', product_id=product_id, workshop_id=workshop_id))

    # Получаем продукт для сообщения
    product = Product.query.get(product_id)

    # Создаем движение (приход)
    movement = Movement(
        product_id=product_id,
        workshop_id=workshop_id,
        quantity=quantity,  # положительное
        movement_type='income',
        reason=reason,
        comment=comment
    )

    db.session.add(movement)

    # Обновляем остаток
    stock = Stock.query.filter_by(product_id=product_id, workshop_id=workshop_id).first()
    if stock:
        stock.quantity += quantity
    else:
        stock = Stock(product_id=product_id, workshop_id=workshop_id, quantity=quantity)
        db.session.add(stock)

    db.session.commit()

    flash(f'✅ Приход: +{quantity} {product.unit}', 'success')
    return redirect(url_for('workshop', workshop_id=workshop_id))


# Обработка расхода
@app.route('/expense', methods=['POST'])
def expense():
    product_id = request.form.get('product_id')
    workshop_id = request.form.get('workshop_id')
    quantity = request.form.get('quantity', type=float)
    reason = request.form.get('reason', 'списание')
    comment = request.form.get('comment', '')

    if quantity <= 0:
        flash('Количество должно быть положительным', 'error')
        return redirect(url_for('movement', product_id=product_id, workshop_id=workshop_id))

    # Проверяем, достаточно ли продуктов
    stock = Stock.query.filter_by(product_id=product_id, workshop_id=workshop_id).first()
    if not stock or stock.quantity < quantity:
        flash(f'❌ Недостаточно продуктов! Доступно: {stock.quantity if stock else 0}', 'error')
        return redirect(url_for('movement', product_id=product_id, workshop_id=workshop_id))

    # Получаем продукт для сообщения
    product = Product.query.get(product_id)

    # Создаем движение (расход)
    movement = Movement(
        product_id=product_id,
        workshop_id=workshop_id,
        quantity=-quantity,  # отрицательное
        movement_type='expense',
        reason=reason,
        comment=comment
    )

    db.session.add(movement)

    # Обновляем остаток
    stock.quantity -= quantity
    db.session.commit()

    flash(f'➖ Расход: -{quantity} {product.unit}', 'success')
    return redirect(url_for('workshop', workshop_id=workshop_id))


# Журнал движений
@app.route('/movements')
def movements():
    workshop_id = request.args.get('workshop', type=int)
    product_id = request.args.get('product', type=int)

    query = Movement.query.order_by(Movement.timestamp.desc())

    if workshop_id:
        query = query.filter_by(workshop_id=workshop_id)
    if product_id:
        query = query.filter_by(product_id=product_id)

    movements = query.all()

    # Для фильтров
    workshops = Workshop.query.all()
    products = Product.query.all()

    return render_template('movements.html',
                           movements=movements,
                           workshops=workshops,
                           products=products,
                           selected_workshop=workshop_id,
                           selected_product=product_id)


# Страница полной инвентаризации
@app.route('/inventory')
def inventory():
    # Получаем все продукты, сортируем по имени
    products = Product.query.order_by(Product.name).all()
    workshops = Workshop.query.all()

    # Группируем продукты по типу
    ingredients = []  # ингредиенты
    semi_finished = []  # полуфабрикаты
    ready_dishes = []  # готовые блюда
    retail = []  # привозные

    for product in products:
        # Для каждого продукта собираем остатки по всем цехам
        stocks = Stock.query.filter_by(product_id=product.id).all()
        # Создаем словарь: ключ - workshop_id, значение - объект Stock
        stocks_dict = {}
        for stock in stocks:
            stocks_dict[stock.workshop_id] = stock

        product_data = {
            'id': product.id,
            'name': product.name,
            'unit': product.unit,
            'category': product.category,
            'type': product.type,
            'stocks': stocks_dict
        }

        if product.type == 'ingredient':
            ingredients.append(product_data)
        elif product.type == 'semi_finished':
            semi_finished.append(product_data)
        elif product.type == 'ready_dish':
            ready_dishes.append(product_data)
        elif product.type == 'retail':
            retail.append(product_data)
        else:
            # Если тип не указан, считаем ингредиентом
            ingredients.append(product_data)

    return render_template('inventory.html',
                           workshops=workshops,
                           ingredients=ingredients,
                           semi_finished=semi_finished,
                           ready_dishes=ready_dishes,
                           retail=retail)


# Сохранение результатов инвентаризации
@app.route('/save_inventory', methods=['POST'])
def save_inventory():
    data = request.json
    updates = data.get('updates', [])

    for update in updates:
        product_id = update.get('product_id')
        workshop_id = update.get('workshop_id')
        quantity = update.get('quantity', type=float)

        if quantity is not None:
            stock = Stock.query.filter_by(
                product_id=product_id,
                workshop_id=workshop_id
            ).first()

            old_quantity = stock.quantity if stock else 0

            if stock:
                stock.quantity = quantity
            else:
                # Если нет записи, создаем новую
                stock = Stock(
                    product_id=product_id,
                    workshop_id=workshop_id,
                    quantity=quantity
                )
                db.session.add(stock)

            # Записываем в историю движений
            if quantity != old_quantity:
                movement = Movement(
                    product_id=product_id,
                    workshop_id=workshop_id,
                    quantity=quantity - old_quantity,
                    movement_type='inventory',
                    reason='инвентаризация',
                    comment=f'Инвентаризация: {old_quantity} → {quantity}'
                )
                db.session.add(movement)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Инвентаризация сохранена'})


# Установка минимального остатка
@app.route('/set_min_stock', methods=['POST'])
def set_min_stock():
    product_id = request.form.get('product_id')
    workshop_id = request.form.get('workshop_id')
    min_stock = request.form.get('min_stock', type=float)

    stock = Stock.query.filter_by(
        product_id=product_id,
        workshop_id=workshop_id
    ).first()

    if stock:
        stock.min_stock = min_stock
        flash(f'Минимальный остаток установлен: {min_stock}', 'success')
    else:
        flash('Продукт не найден в этом цехе', 'error')

    db.session.commit()
    return redirect(url_for('workshop', workshop_id=workshop_id))


# API для получения статистики по минимальным остаткам
@app.route('/api/low_stock_count')
def low_stock_count():
    count = Stock.query.filter(Stock.quantity < Stock.min_stock).count()
    return jsonify({'count': count})


# Страница заказов (ручное создание)
@app.route('/orders')
def orders():
    # Получаем активные заказы
    active_orders = Order.query.filter_by(status='active').order_by(Order.priority.desc(), Order.created_at).all()
    # Получаем выполненные заказы (последние 20)
    completed_orders = Order.query.filter_by(status='completed').order_by(Order.completed_at.desc()).limit(20).all()

    workshops = Workshop.query.all()
    products = Product.query.all()

    return render_template('orders.html',
                           active_orders=active_orders,
                           completed_orders=completed_orders,
                           workshops=workshops,
                           products=products)


# Создание нового заказа
@app.route('/create_order', methods=['POST'])
def create_order():
    product_id = request.form.get('product_id')
    workshop_id = request.form.get('workshop_id')
    quantity = request.form.get('quantity', type=float)
    priority = request.form.get('priority', 'normal')
    comment = request.form.get('comment', '')

    product = Product.query.get(product_id)

    order = Order(
        product_id=product_id,
        workshop_id=workshop_id,
        quantity=quantity,
        unit=product.unit,
        priority=priority,
        comment=comment,
        status='active'
    )

    db.session.add(order)
    db.session.commit()

    flash(f'✅ Заказ создан: {product.name} {quantity} {product.unit}', 'success')
    return redirect(url_for('orders'))


# Отметить заказ как выполненный
@app.route('/complete_order/<int:order_id>')
def complete_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = 'completed'
    order.completed_at = datetime.utcnow()
    db.session.commit()

    flash(f'✅ Заказ выполнен: {order.product.name}', 'success')
    return redirect(url_for('orders'))


# Отмена заказа
@app.route('/cancel_order/<int:order_id>')
def cancel_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = 'cancelled'
    db.session.commit()

    flash(f'❌ Заказ отменён: {order.product.name}', 'info')
    return redirect(url_for('orders'))


# Редактирование заказа
@app.route('/edit_order/<int:order_id>', methods=['POST'])
def edit_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.quantity = request.form.get('quantity', type=float)
    order.priority = request.form.get('priority', 'normal')
    order.comment = request.form.get('comment', '')
    db.session.commit()

    flash(f'✏️ Заказ обновлён', 'success')
    return redirect(url_for('orders'))


# Быстрое создание заказа из инвентаризации
@app.route('/quick_order/<int:product_id>/<int:workshop_id>')
def quick_order(product_id, workshop_id):
    product = Product.query.get(product_id)
    # Проверяем, есть ли уже активный заказ на этот продукт
    existing = Order.query.filter_by(
        product_id=product_id,
        workshop_id=workshop_id,
        status='active'
    ).first()

    if existing:
        flash(f'⚠️ Заказ на {product.name} уже существует', 'warning')
    else:
        order = Order(
            product_id=product_id,
            workshop_id=workshop_id,
            quantity=1.0,
            unit=product.unit,
            priority='normal',
            status='active'
        )
        db.session.add(order)
        db.session.commit()
        flash(f'📋 Быстрый заказ создан: {product.name}', 'success')

    return redirect(url_for('orders'))


# ========== ТЕХНОЛОГИЧЕСКИЕ КАРТЫ (РЕЦЕПТЫ) ==========

# Список всех рецептов
@app.route('/recipes')
def recipes():
    recipes = Recipe.query.order_by(Recipe.name).all()
    return render_template('recipes.html', recipes=recipes)


# Просмотр одного рецепта
@app.route('/recipe/<int:recipe_id>')
def recipe_view(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)
    return render_template('recipe_view.html', recipe=recipe)


# Создание нового рецепта
@app.route('/recipe/create', methods=['GET', 'POST'])
def recipe_create():
    if request.method == 'POST':
        name = request.form.get('name')
        dish_id = request.form.get('dish_id')
        workshop_id = request.form.get('workshop_id')
        portions = request.form.get('portions', 1, type=int)
        cooking_time = request.form.get('cooking_time', type=int)
        instructions = request.form.get('instructions', '')

        recipe = Recipe(
            name=name,
            dish_id=dish_id,
            workshop_id=workshop_id,
            portions=portions,
            cooking_time=cooking_time,
            instructions=instructions
        )
        db.session.add(recipe)
        db.session.commit()

        flash(f'✅ Рецепт "{name}" создан. Теперь добавьте ингредиенты.', 'success')
        return redirect(url_for('recipe_edit', recipe_id=recipe.id))

    # GET запрос - показываем форму
    dishes = Product.query.filter_by(type='ready_dish').all()
    workshops = Workshop.query.all()
    return render_template('recipe_create.html', dishes=dishes, workshops=workshops)


# Редактирование рецепта
@app.route('/recipe/<int:recipe_id>/edit', methods=['GET', 'POST'])
def recipe_edit(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)

    if request.method == 'POST':
        recipe.name = request.form.get('name')
        recipe.portions = request.form.get('portions', 1, type=int)
        recipe.cooking_time = request.form.get('cooking_time', type=int)
        recipe.instructions = request.form.get('instructions', '')
        db.session.commit()
        flash('✅ Рецепт обновлён', 'success')
        return redirect(url_for('recipe_view', recipe_id=recipe.id))

    # GET запрос - показываем форму
    workshops = Workshop.query.all()
    products = Product.query.filter_by(type='ingredient').all()
    return render_template('recipe_edit.html', recipe=recipe, workshops=workshops, products=products)


# Добавление ингредиента в рецепт
@app.route('/recipe/<int:recipe_id>/add_ingredient', methods=['POST'])
def recipe_add_ingredient(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)

    product_id = request.form.get('product_id')
    quantity = request.form.get('quantity', type=float)
    unit = request.form.get('unit')
    notes = request.form.get('notes', '')

    product = Product.query.get(product_id)

    ingredient = RecipeIngredient(
        recipe_id=recipe_id,
        product_id=product_id,
        quantity=quantity,
        unit=unit or product.unit,
        notes=notes
    )
    db.session.add(ingredient)
    db.session.commit()

    flash(f'✅ Ингредиент {product.name} добавлен', 'success')
    return redirect(url_for('recipe_edit', recipe_id=recipe_id))


# Удаление ингредиента из рецепта
@app.route('/recipe/<int:recipe_id>/remove_ingredient/<int:ingredient_id>')
def recipe_remove_ingredient(recipe_id, ingredient_id):
    ingredient = RecipeIngredient.query.get_or_404(ingredient_id)
    db.session.delete(ingredient)
    db.session.commit()
    flash('✅ Ингредиент удалён', 'success')
    return redirect(url_for('recipe_edit', recipe_id=recipe_id))


# Приготовление блюда (списание ингредиентов)
@app.route('/recipe/<int:recipe_id>/cook', methods=['POST'])
def recipe_cook(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)
    portions = request.form.get('portions', 1, type=int)

    if portions <= 0:
        flash('❌ Количество порций должно быть больше 0', 'error')
        return redirect(url_for('recipe_view', recipe_id=recipe_id))

    # Проверяем, хватает ли ингредиентов
    missing = []
    for ingredient in recipe.ingredients:
        total_needed = ingredient.quantity * portions
        # Ищем продукт в нужном цехе
        stock = Stock.query.filter_by(
            product_id=ingredient.product_id,
            workshop_id=recipe.workshop_id
        ).first()

        if not stock or stock.quantity < total_needed:
            missing.append({
                'name': ingredient.product.name,
                'needed': total_needed,
                'available': stock.quantity if stock else 0,
                'unit': ingredient.unit
            })

    if missing:
        # Не хватает продуктов
        flash('❌ Недостаточно продуктов:', 'error')
        for item in missing:
            flash(f'  • {item["name"]}: нужно {item["needed"]} {item["unit"]}, есть {item["available"]} {item["unit"]}',
                  'error')
        return redirect(url_for('recipe_view', recipe_id=recipe_id))

    # Всё хорошо - списываем ингредиенты
    for ingredient in recipe.ingredients:
        total_needed = ingredient.quantity * portions
        stock = Stock.query.filter_by(
            product_id=ingredient.product_id,
            workshop_id=recipe.workshop_id
        ).first()

        if stock:
            old_quantity = stock.quantity
            stock.quantity -= total_needed

            # Записываем в историю движений
            movement = Movement(
                product_id=ingredient.product_id,
                workshop_id=recipe.workshop_id,
                quantity=-total_needed,
                movement_type='expense',
                reason='приготовление',
                comment=f'Приготовлено {portions} порций "{recipe.name}"'
            )
            db.session.add(movement)

    db.session.commit()
    flash(f'✅ Приготовлено {portions} порций "{recipe.name}"', 'success')
    return redirect(url_for('recipe_view', recipe_id=recipe_id))


# ========== ОТЧЁТЫ ==========
import pandas as pd
from flask import send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta


@app.route('/reports')
def reports():
    """Страница с отчётами"""
    workshops = Workshop.query.all()
    low_stocks_count = Stock.query.filter(Stock.quantity < Stock.min_stock).count()
    active_orders_count = Order.query.filter_by(status='active').count()

    # Для каждого цеха считаем количество продуктов
    workshop_stats = []
    for workshop in workshops:
        stocks_count = Stock.query.filter_by(workshop_id=workshop.id).count()
        workshop_stats.append({
            'name': workshop.name,
            'stocks_count': stocks_count
        })

    return render_template('reports.html',
                           workshops=workshops,
                           workshop_stats=workshop_stats,
                           low_stocks_count=low_stocks_count,
                           active_orders_count=active_orders_count)


@app.route('/export_excel')
def export_excel():
    """Полный отчёт в Excel"""
    # Создаём книгу Excel
    wb = Workbook()

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== ЛИСТ 1: Остатки по цехам ==========
    ws1 = wb.active
    ws1.title = "Остатки по цехам"

    # Заголовки
    headers = ['Цех', 'Продукт', 'Категория', 'Количество', 'Ед. изм.', 'Тип', 'Минимум', 'Статус']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Данные
    stocks = Stock.query.all()
    row = 2
    for stock in stocks:
        ws1.cell(row=row, column=1, value=stock.workshop.name).border = border
        ws1.cell(row=row, column=2, value=stock.product.name).border = border
        ws1.cell(row=row, column=3, value=stock.product.category or '-').border = border
        ws1.cell(row=row, column=4, value=stock.quantity).border = border
        ws1.cell(row=row, column=5, value=stock.product.unit).border = border
        ws1.cell(row=row, column=6, value=stock.product.type).border = border
        ws1.cell(row=row, column=7, value=stock.min_stock).border = border

        # Статус (норма/мало)
        if stock.min_stock > 0 and stock.quantity < stock.min_stock:
            status = "⚠️ МАЛО"
            ws1.cell(row=row, column=8, value=status).border = border
            ws1.cell(row=row, column=8).font = Font(color="FF0000")
        else:
            status = "✓ норма"
            ws1.cell(row=row, column=8, value=status).border = border

        row += 1

    # Автоширина колонок
    for col in range(1, 9):
        ws1.column_dimensions[chr(64 + col)].width = 15

    # ========== ЛИСТ 2: Сводка по цехам ==========
    ws2 = wb.create_sheet("Сводка по цехам")

    # Заголовки
    summary_headers = ['Цех', 'Количество позиций', 'Общее количество', 'Продуктов ниже минимума']
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Данные
    workshops = Workshop.query.all()
    row = 2
    for workshop in workshops:
        workshop_stocks = Stock.query.filter_by(workshop_id=workshop.id).all()
        low_stocks = [s for s in workshop_stocks if s.min_stock > 0 and s.quantity < s.min_stock]

        ws2.cell(row=row, column=1, value=workshop.name).border = border
        ws2.cell(row=row, column=2, value=len(workshop_stocks)).border = border
        ws2.cell(row=row, column=3, value=sum(s.quantity for s in workshop_stocks)).border = border
        ws2.cell(row=row, column=4, value=len(low_stocks)).border = border
        row += 1

    # ========== ЛИСТ 3: Продукты ниже минимума ==========
    ws3 = wb.create_sheet("Нужно заказать")

    # Заголовки
    low_headers = ['Цех', 'Продукт', 'Текущий остаток', 'Минимум', 'Нужно заказать', 'Ед. изм.']
    for col, header in enumerate(low_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border

    # Данные
    low_stocks = Stock.query.filter(Stock.quantity < Stock.min_stock).all()
    row = 2
    for stock in low_stocks:
        if stock.min_stock > 0:
            ws3.cell(row=row, column=1, value=stock.workshop.name).border = border
            ws3.cell(row=row, column=2, value=stock.product.name).border = border
            ws3.cell(row=row, column=3, value=stock.quantity).border = border
            ws3.cell(row=row, column=4, value=stock.min_stock).border = border
            ws3.cell(row=row, column=5, value=stock.min_stock - stock.quantity).border = border
            ws3.cell(row=row, column=5).font = Font(color="FF0000", bold=True)
            ws3.cell(row=row, column=6, value=stock.product.unit).border = border
            row += 1

    # ========== ЛИСТ 4: Активные заказы ==========
    ws4 = wb.create_sheet("Активные заказы")

    # Заголовки
    order_headers = ['Цех', 'Продукт', 'Количество', 'Ед. изм.', 'Приоритет', 'Дата создания']
    for col, header in enumerate(order_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border

    # Данные
    active_orders = Order.query.filter_by(status='active').all()
    row = 2
    for order in active_orders:
        priority_map = {'high': '🔴 Высокий', 'normal': '🟡 Средний', 'low': '🟢 Низкий'}
        ws4.cell(row=row, column=1, value=order.workshop.name).border = border
        ws4.cell(row=row, column=2, value=order.product.name).border = border
        ws4.cell(row=row, column=3, value=order.quantity).border = border
        ws4.cell(row=row, column=4, value=order.unit).border = border
        ws4.cell(row=row, column=5, value=priority_map.get(order.priority, order.priority)).border = border
        ws4.cell(row=row, column=6, value=order.created_at.strftime('%d.%m.%Y %H:%M')).border = border
        row += 1

    # ========== ЛИСТ 5: Движения за последние 7 дней ==========
    ws5 = wb.create_sheet("Движения (7 дней)")

    # Заголовки
    move_headers = ['Дата', 'Цех', 'Продукт', 'Тип', 'Количество', 'Ед. изм.', 'Причина', 'Комментарий']
    for col, header in enumerate(move_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border

    # Данные за последние 7 дней
    week_ago = datetime.utcnow() - timedelta(days=7)
    recent_movements = Movement.query.filter(Movement.timestamp >= week_ago).order_by(Movement.timestamp.desc()).all()
    row = 2
    for move in recent_movements:
        move_type = "📦 Приход" if move.quantity > 0 else "➖ Расход"
        color = "28A745" if move.quantity > 0 else "DC3545"

        ws5.cell(row=row, column=1, value=move.timestamp.strftime('%d.%m.%Y %H:%M')).border = border
        ws5.cell(row=row, column=2, value=move.workshop.name).border = border
        ws5.cell(row=row, column=3, value=move.product.name).border = border
        ws5.cell(row=row, column=4, value=move_type).border = border
        ws5.cell(row=row, column=4).font = Font(color=color)
        ws5.cell(row=row, column=5, value=abs(move.quantity)).border = border
        ws5.cell(row=row, column=6, value=move.product.unit).border = border
        ws5.cell(row=row, column=7, value=move.reason or '-').border = border
        ws5.cell(row=row, column=8, value=move.comment or '-').border = border
        row += 1

    # Сохраняем в память
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Отправляем файл
    filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_movements')
def export_movements():
    """Отчёт только по движениям"""
    days = request.args.get('days', 30, type=int)

    # Создаём DataFrame с движениями
    movements = Movement.query.order_by(Movement.timestamp.desc()).limit(1000).all()

    data = []
    for m in movements:
        data.append({
            'Дата': m.timestamp.strftime('%d.%m.%Y %H:%M'),
            'Цех': m.workshop.name,
            'Продукт': m.product.name,
            'Тип': 'Приход' if m.quantity > 0 else 'Расход',
            'Количество': abs(m.quantity),
            'Ед.': m.product.unit,
            'Причина': m.reason or '-',
            'Комментарий': m.comment or '-'
        })

    df = pd.DataFrame(data)

    # Сохраняем в Excel
    excel_file = io.BytesIO()
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Движения')

    excel_file.seek(0)

    filename = f"movements_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ========== РЕЗЕРВНОЕ КОПИРОВАНИЕ ==========
from backup import BackupManager

backup_manager = BackupManager()


@app.route('/backups')
def backups():
    """Страница управления бэкапами"""
    backups_list = backup_manager.list_backups()
    return render_template('backups.html', backups=backups_list)


@app.route('/create_backup', methods=['POST'])
def create_backup():
    """Создание нового бэкапа"""
    comment = request.form.get('comment', '')
    try:
        path = backup_manager.create_backup(comment=comment)
        flash(f'✅ Бэкап успешно создан: {os.path.basename(path)}', 'success')
    except Exception as e:
        flash(f'❌ Ошибка при создании бэкапа: {str(e)}', 'error')
    return redirect(url_for('backups'))


@app.route('/restore_backup/<filename>')
def restore_backup(filename):
    """Восстановление из бэкапа"""
    try:
        if backup_manager.restore_backup(filename):
            flash(f'✅ База данных восстановлена из {filename}', 'success')
        else:
            flash(f'❌ Не удалось восстановить из {filename}', 'error')
    except Exception as e:
        flash(f'❌ Ошибка при восстановлении: {str(e)}', 'error')
    return redirect(url_for('backups'))


@app.route('/download_backup/<filename>')
def download_backup(filename):
    """Скачивание бэкапа"""
    from flask import send_file
    backup_path = os.path.join('backups', filename)
    return send_file(backup_path, as_attachment=True)


@app.route('/delete_backup/<filename>')
def delete_backup(filename):
    """Удаление бэкапа"""
    try:
        os.remove(os.path.join('backups', filename))
        flash(f'✅ Бэкап {filename} удалён', 'success')
    except Exception as e:
        flash(f'❌ Ошибка при удалении: {str(e)}', 'error')
    return redirect(url_for('backups'))


# ========== ОБЪЕДИНЕНИЕ ПРОДУКТОВ ==========
@app.route('/merge_products', methods=['POST'])
def merge_products():
    """Объединение дублирующихся продуктов"""
    product_id_1 = request.form.get('product_id_1', type=int)
    product_id_2 = request.form.get('product_id_2', type=int)

    product1 = Product.query.get(product_id_1)
    product2 = Product.query.get(product_id_2)

    if not product1 or not product2:
        flash('❌ Продукты не найдены', 'error')
        return redirect(url_for('inventory'))

    # Переносим все остатки из product2 в product1
    stocks_to_move = Stock.query.filter_by(product_id=product_id_2).all()
    for stock in stocks_to_move:
        # Ищем существующий остаток product1 в том же цехе
        existing_stock = Stock.query.filter_by(
            product_id=product_id_1,
            workshop_id=stock.workshop_id
        ).first()

        if existing_stock:
            # Складываем количества
            existing_stock.quantity += stock.quantity
            db.session.delete(stock)
        else:
            # Перенаправляем связь
            stock.product_id = product_id_1

    # Переносим заказы
    Order.query.filter_by(product_id=product_id_2).update({'product_id': product_id_1})

    # Переносим движения
    Movement.query.filter_by(product_id=product_id_2).update({'product_id': product_id_1})

    # Переносим рецепты
    RecipeIngredient.query.filter_by(product_id=product_id_2).update({'product_id': product_id_1})

    # Удаляем дубликат
    db.session.delete(product2)
    db.session.commit()

    flash(f'✅ Продукты объединены: {product1.name} + {product2.name}', 'success')
    return redirect(url_for('inventory'))


# Дашборд (главная страница с аналитикой)
@app.route('/dashboard')
def dashboard():
    # Общая статистика
    total_products = Product.query.count()
    total_stock = Stock.query.count()
    low_stock_count = Stock.query.filter(Stock.quantity < Stock.min_stock).count()
    active_orders = Order.query.filter_by(status='active').count()

    # Последние движения (10 штук)
    recent_movements = Movement.query.order_by(Movement.timestamp.desc()).limit(10).all()

    # Продукты, которые заканчиваются (топ 5)
    low_stock_products = []
    low_stocks = Stock.query.filter(Stock.quantity < Stock.min_stock).limit(5).all()
    for stock in low_stocks:
        low_stock_products.append({
            'name': stock.product.name,
            'workshop': stock.workshop.name,
            'current': stock.quantity,
            'min': stock.min_stock,
            'unit': stock.product.unit
        })

    # Данные для графика (последние 7 дней)
    from datetime import datetime, timedelta
    movements_data = []
    for i in range(6, -1, -1):
        day = datetime.now() - timedelta(days=i)
        day_start = datetime(day.year, day.month, day.day)
        day_end = day_start + timedelta(days=1)

        day_movements = Movement.query.filter(
            Movement.timestamp >= day_start,
            Movement.timestamp < day_end
        ).count()

        # Высота столбца (макс 150px)
        height = min(150, day_movements * 30)

        movements_data.append({
            'label': day.strftime('%d.%m'),
            'height': height,
            'count': day_movements
        })

    return render_template('dashboard.html',
                           total_products=total_products,
                           total_stock=total_stock,
                           low_stock_count=low_stock_count,
                           active_orders=active_orders,
                           recent_movements=recent_movements,
                           low_stock_products=low_stock_products,
                           movements=movements_data)
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)


